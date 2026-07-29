"""Ingest EIA Form 860 Schedule 3–1 generators into ``generation_units`` (batteries excluded).

Run from repo root::

    DATABASE_URL=... py -3 -m app.scripts.ingest_eia_860

Environment: ``DATABASE_URL`` (PostgreSQL + PostGIS).
"""

from __future__ import annotations

import io
import logging
import os
import re
import sys
import zipfile
from collections import Counter
from datetime import date, datetime
from typing import Any, Callable, Iterator
from urllib.parse import urljoin

import openpyxl
import requests
from bs4 import BeautifulSoup
from psycopg2.extras import execute_batch

from app.scripts._db import connect
from app.services.iso_lookup import iso_for_point

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

USER_AGENT = "GridAlpha/1.0 (+https://github.com/aquilesguerretta/gridalpha-v2)"
EIA_860_INDEX = "https://www.eia.gov/electricity/data/eia860/"

FUEL_CODE_MAP: dict[str, str] = {
    "NG": "gas",
    "NGCC": "gas",
    "NGCT": "gas",
    "BIT": "coal",
    "SUB": "coal",
    "LIG": "coal",
    "RC": "coal",
    "WC": "coal",
    "NUC": "nuclear",
    "WND": "wind",
    "SUN": "solar",
    "WAT": "hydro",
    "PS": "pumped",
    "WDS": "biomass",
    "MSW": "biomass",
    "OBG": "biomass",
    "LFG": "biomass",
    "WOO": "biomass",
    "GEO": "geothermal",
    "DFO": "oil",
    "RFO": "oil",
    "KER": "oil",
}


def _norm_header_val(val: Any) -> str:
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val)).strip().lower()


def _col_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    """Map canonical keys to 1-based column indices."""
    idx: dict[str, int] = {}
    for i, val in enumerate(header_row, start=1):
        h = _norm_header_val(val)
        if not h:
            continue
        if h in ("plant id", "plant code"):
            idx.setdefault("plant_id", i)
        elif "generator id" in h:
            idx.setdefault("generator_id", i)
        elif h == "technology":
            idx.setdefault("technology", i)
        elif h in (
            "energy source 1",
            "energy source 1 code",
            "energy source",
        ):
            idx.setdefault("energy_source_1", i)
        elif "nameplate capacity" in h and "mw" in h:
            idx.setdefault("capacity", i)
        elif "net summer capacity" in h and "mw" in h:
            idx.setdefault("summer_capacity", i)
        elif h == "state":
            idx.setdefault("state", i)
        elif h in ("latitude",):
            idx.setdefault("lat", i)
        elif h in ("longitude",):
            idx.setdefault("lon", i)
        elif h == "plant name":
            idx.setdefault("plant_name", i)
        elif h == "generator name":
            idx.setdefault("generator_name", i)
        elif "utility name" in h or h == "entity name":
            idx.setdefault("owner", i)
        elif "operating month" in h and "year" in h:
            idx.setdefault("operating_date", i)
        elif h == "operating year":
            idx.setdefault("operating_date", i)
        elif "retirement" in h and "month" in h:
            idx.setdefault("retirement_date", i)
        elif "status" in h or "service" in h:
            idx.setdefault("status_raw", i)
    if "capacity" not in idx and "summer_capacity" in idx:
        idx["capacity"] = idx["summer_capacity"]
    return idx


def scan_header(
    rows: Iterator[tuple[Any, ...]],
    required: tuple[str, ...],
    col_map_fn: Callable[[tuple[Any, ...]], dict[str, int]] = _col_map,
    max_scan: int = 8,
) -> dict[str, int] | None:
    """Advance ``rows`` past EIA's title banner and return the real column map.

    EIA workbooks open with one or more banner rows before the header, so the
    first row is not the header. Consumes rows until one maps every key in
    ``required``; the iterator is then positioned on the first data row.
    """
    for _ in range(max_scan):
        try:
            row = next(rows)
        except StopIteration:
            return None
        if not row:
            continue
        cmap = col_map_fn(tuple(row))
        if all(k in cmap for k in required):
            return cmap
    return None


def _cell_str(vals: tuple[Any, ...], col: int | None) -> str:
    if col is None or col < 1 or col > len(vals):
        return ""
    v = vals[col - 1]
    if v is None:
        return ""
    return str(v).strip()


def _cell_float(vals: tuple[Any, ...], col: int | None) -> float | None:
    s = _cell_str(vals, col)
    if not s:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def _parse_date(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt).date()
        except ValueError:
            continue
    m = re.search(r"(\d{4})", s)
    if m:
        try:
            return date(int(m.group(1)), 1, 1)
        except ValueError:
            pass
    return None


def _is_battery(tech: str, es1: str) -> bool:
    t = (tech or "").upper()
    e = (es1 or "").upper().strip()
    return "BATTERY" in t or "BATTERIES" in t or e in ("MWH", "BAT")


def _map_fuel(technology: str, energy_source_1: str) -> str:
    e = (energy_source_1 or "").upper().strip()
    if e in FUEL_CODE_MAP:
        return FUEL_CODE_MAP[e]
    tok = re.sub(r"[^A-Z0-9]", "", (technology or "").upper())
    for code, fuel in FUEL_CODE_MAP.items():
        if code in tok:
            return fuel
    words = (technology or "").lower()
    if "wind" in words:
        return "wind"
    if "solar" in words or "photovoltaic" in words:
        return "solar"
    if "hydro" in words or "water" in words:
        if "pump" in words:
            return "pumped"
        return "hydro"
    if "nuclear" in words:
        return "nuclear"
    if "coal" in words:
        return "coal"
    if "gas" in words or "methane" in words:
        return "gas"
    if "oil" in words or "diesel" in words or "kerosene" in words:
        return "oil"
    if "bio" in words or "biomass" in words or "waste" in words:
        return "biomass"
    if "geothermal" in words:
        return "geothermal"
    return "other"


def _map_proposed_status(status_cell: str) -> str:
    s = (status_cell or "").upper()
    if "(CN)" in s or " CANCELLED" in s or "CANCEL" in s:
        return "cancelled"
    if "(PL)" in s or "PLANNED" in s:
        return "planned"
    if "(SB)" in s or "STANDBY" in s:
        return "standby"
    if any(
        x in s for x in ("(T)", "(U)", "(V)", "UNDER CONSTRUCTION", "UNDER_CONSTRUCTION")
    ):
        return "under-construction"
    return "planned"


def _operable_status(_status_cell: str) -> str:
    return "operating"


def _retired_status(_status_cell: str) -> str:
    return "retired"


# EIA publishes the annual archive as eia860YYYY.zip, plus an eia860YYYYER.zip
# "early release" ahead of the final year. The pre-2026 f860YYYY.zip name is gone.
_ZIP_RE = re.compile(r"eia860(20\d{2})(ER)?\.zip$", re.I)


def resolve_latest_860_zip() -> tuple[str, int]:
    """Return (url, year) for the newest annual 860 archive.

    ``EIA_860_ZIP_URL`` pins a specific archive (e.g. to prefer the validated
    final year over a preliminary early release).
    """
    override = os.environ.get("EIA_860_ZIP_URL", "").strip()
    if override:
        m = _ZIP_RE.search(override)
        return override, int(m.group(1)) if m else 0

    resp = requests.get(EIA_860_INDEX, headers={"User-Agent": USER_AGENT}, timeout=120)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")
    best_url: str | None = None
    best_key: tuple[int, int] = (-1, -1)
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "/archive/" in href.lower():
            continue
        m = _ZIP_RE.search(href)
        if not m:
            continue
        # Same year: the final archive outranks the early release.
        key = (int(m.group(1)), 0 if m.group(2) else 1)
        if key <= best_key:
            continue
        best_key = key
        best_url = href if href.startswith("http") else urljoin(EIA_860_INDEX, href)
    if not best_url:
        raise RuntimeError("could not resolve latest eia860YYYY.zip from EIA 860 index")
    return best_url, best_key[0]


def find_zip_entry(zf: zipfile.ZipFile, token: str) -> str | None:
    """Locate a workbook by name fragment — early-release entries carry a suffix."""
    for name in zf.namelist():
        if token in name and name.lower().endswith(".xlsx"):
            return name
    return None


def load_plant_coords(zf: zipfile.ZipFile) -> dict[str, tuple[float, float]]:
    """Map plant code -> (lat, lon) from schedule 2.

    Schedule 3_1 (generators) carries no coordinates; they live only in the
    plant file, keyed by plant code.
    """
    entry = find_zip_entry(zf, "2___Plant")
    if not entry:
        raise RuntimeError("could not find 2___Plant workbook in EIA 860 zip")
    wb = openpyxl.load_workbook(io.BytesIO(zf.read(entry)), read_only=True, data_only=True)
    coords: dict[str, tuple[float, float]] = {}
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        cmap = scan_header(rows, ("plant_id", "lat", "lon"))
        if not cmap:
            raise RuntimeError(f"could not locate header row in {entry}")
        for raw in rows:
            if not raw:
                continue
            row = tuple(raw)
            code = _plant_key(_cell_str(row, cmap.get("plant_id")))
            lat = _cell_float(row, cmap.get("lat"))
            lon = _cell_float(row, cmap.get("lon"))
            if not code or lat is None or lon is None:
                continue
            coords[code] = (lat, lon)
    finally:
        wb.close()
    logger.info("loaded %s plant coordinates from %s", len(coords), entry)
    return coords


def _plant_key(raw: str) -> str:
    """Normalize a plant code — the two workbooks type it as int and as float."""
    s = (raw or "").strip()
    if not s:
        return ""
    try:
        return str(int(float(s)))
    except ValueError:
        return s


def download_zip(url: str) -> bytes:
    r = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=600)
    r.raise_for_status()
    return r.content


def build_rows(
    wb: openpyxl.workbook.workbook.Workbook,
    plant_coords: dict[str, tuple[float, float]],
) -> list[tuple]:
    handlers: list[tuple[str, Any]] = [
        ("Operable", _operable_status),
        ("Proposed", _map_proposed_status),
    ]
    for retired_sheet in (
        "Retired and Canceled",
        "Retired",
        "Retired and Cancelled",
    ):
        if retired_sheet in wb.sheetnames:
            handlers.append((retired_sheet, _retired_status))
            break

    by_id: dict[str, tuple] = {}
    for sheet_name, status_fn in handlers:
        if sheet_name not in wb.sheetnames:
            logger.warning("sheet %s not found — skip", sheet_name)
            continue
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        need = ("plant_id", "generator_id", "state")
        cmap = scan_header(it, need)
        if not cmap:
            logger.warning("sheet %s missing required columns %s", sheet_name, need)
            continue
        for raw in it:
            if not raw:
                continue
            row = tuple(raw)
            plant_id = _cell_str(row, cmap.get("plant_id"))
            gen_id = _cell_str(row, cmap.get("generator_id"))
            if not plant_id or not gen_id:
                continue
            technology = _cell_str(row, cmap.get("technology"))
            es1 = _cell_str(row, cmap.get("energy_source_1"))
            if _is_battery(technology, es1):
                continue
            cap = _cell_float(row, cmap.get("capacity"))
            if cap is None or cap <= 0:
                continue
            state = _cell_str(row, cmap.get("state")).upper()[:2] or "ZZ"
            coord = plant_coords.get(_plant_key(plant_id))
            if coord is None:
                continue
            lat, lon = coord
            fuel = _map_fuel(technology, es1)
            st_raw = _cell_str(row, cmap.get("status_raw"))
            status = status_fn(st_raw)
            plant_name = _cell_str(row, cmap.get("plant_name")) or f"Plant {plant_id}"
            gen_name = _cell_str(row, cmap.get("generator_name"))
            name = gen_name or plant_name
            owner = _cell_str(row, cmap.get("owner")) or None
            oid = f"eia-{plant_id}-{gen_id}"
            cod: date | None = None
            if cmap.get("operating_date"):
                cod = _parse_date(row[cmap["operating_date"] - 1])
            ret: date | None = None
            if cmap.get("retirement_date"):
                ret = _parse_date(row[cmap["retirement_date"] - 1])
            iso = iso_for_point(lat, lon)
            try:
                eia_plant_int = int(float(plant_id))
            except ValueError:
                eia_plant_int = None
            by_id[oid] = (
                oid,
                eia_plant_int,
                gen_id or None,
                name,
                owner,
                iso,
                state,
                fuel,
                float(cap),
                status,
                cod,
                ret,
                lon,
                lat,
            )
    return list(by_id.values())


UPSERT_SQL = """
INSERT INTO generation_units (
  id, eia_plant_id, eia_generator_id, name, owner, iso, state, fuel,
  capacity_mw, status, cod_date, retirement_date, geom
) VALUES (
  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
  ST_SetSRID(ST_MakePoint(%s, %s), 4326)
)
ON CONFLICT (id) DO UPDATE SET
  eia_plant_id = EXCLUDED.eia_plant_id,
  eia_generator_id = EXCLUDED.eia_generator_id,
  name = EXCLUDED.name,
  owner = EXCLUDED.owner,
  iso = EXCLUDED.iso,
  state = EXCLUDED.state,
  fuel = EXCLUDED.fuel,
  capacity_mw = EXCLUDED.capacity_mw,
  status = EXCLUDED.status,
  cod_date = EXCLUDED.cod_date,
  retirement_date = EXCLUDED.retirement_date,
  geom = EXCLUDED.geom;
"""


def main() -> int:
    url, year = resolve_latest_860_zip()
    logger.info("latest EIA 860 archive: %s (year=%s)", url, year)
    raw = download_zip(url)
    if raw[:2] != b"PK":
        logger.error("download is not a ZIP — got %s", raw[:200])
        return 2
    zf = zipfile.ZipFile(io.BytesIO(raw))
    inner_path = find_zip_entry(zf, "3_1_Generator")
    if not inner_path:
        logger.error("could not find 3_1_Generator workbook in zip")
        return 2
    plant_coords = load_plant_coords(zf)
    data = zf.read(inner_path)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows = build_rows(wb, plant_coords)
    wb.close()
    logger.info("parsed %s generator rows from %s (pre-upsert)", len(rows), inner_path)

    fuel_ct = Counter(r[7] for r in rows)
    iso_ct = Counter(r[5] for r in rows)
    logger.info("fuel distribution: %s", dict(fuel_ct.most_common(12)))
    logger.info("ISO distribution: %s", dict(iso_ct.most_common(12)))

    conn = connect()
    try:
        with conn.cursor() as cur:
            # execute_batch, not executemany: psycopg2 sends one round trip per
            # row for executemany, which over the public proxy costs hours.
            for batch_start in range(0, len(rows), 500):
                batch = rows[batch_start : batch_start + 500]
                execute_batch(cur, UPSERT_SQL, batch, page_size=500)
                conn.commit()
                logger.info("committed %s / %s", min(batch_start + 500, len(rows)), len(rows))
    finally:
        conn.close()
    logger.info("upserted %s generation_units rows", len(rows))
    return 0


if __name__ == "__main__":
    sys.exit(main())
