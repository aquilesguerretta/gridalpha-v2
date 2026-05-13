"""Ingest EIA Form 860M monthly battery storage rows into ``battery_assets``.

Run from repo root::

    DATABASE_URL=... py -3 -m app.scripts.ingest_eia_860m

Environment: ``DATABASE_URL`` (PostgreSQL + PostGIS).
"""

from __future__ import annotations

import io
import logging
import re
import sys
from collections import Counter
from datetime import date, datetime
from typing import Any
from urllib.parse import urljoin

import openpyxl
import requests
from bs4 import BeautifulSoup

from app.scripts._db import connect
from app.scripts.ingest_eia_860 import (  # reuse status + date helpers
    USER_AGENT,
    _map_proposed_status,
    _operable_status,
    _parse_date,
    _retired_status,
)
from app.services.iso_lookup import iso_for_point

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

EIA_860M_INDEX = "https://www.eia.gov/electricity/data/eia860m/"

_MONTH_ORDER = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


def _norm_header_val(val: Any) -> str:
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val)).strip().lower()


def _col_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    idx: dict[str, int] = {}
    for i, val in enumerate(header_row, start=1):
        h = _norm_header_val(val)
        if not h:
            continue
        if h in ("plant id", "plant code"):
            idx.setdefault("plant_id", i)
        elif "generator id" in h:
            idx.setdefault("generator_id", i)
        elif h == "technology":
            idx.setdefault("technology", i)
        elif h in ("energy source", "energy source 1", "energy source 1 code"):
            idx.setdefault("energy_source_1", i)
        elif "net summer capacity" in h and "mw" in h:
            idx.setdefault("capacity_mw", i)
        elif "nameplate energy capacity" in h and "mwh" in h:
            idx.setdefault("capacity_mwh", i)
        elif h == "state":
            idx.setdefault("state", i)
        elif h in ("latitude",):
            idx.setdefault("lat", i)
        elif h in ("longitude",):
            idx.setdefault("lon", i)
        elif h == "plant name":
            idx.setdefault("plant_name", i)
        elif h == "generator name":
            idx.setdefault("generator_name", i)
        elif "utility name" in h or h == "entity name":
            idx.setdefault("owner", i)
        elif "operating month" in h and "year" in h:
            idx.setdefault("operating_date", i)
        elif h == "operating year":
            idx.setdefault("operating_date", i)
        elif "retirement" in h and "month" in h:
            idx.setdefault("retirement_date", i)
        elif "status" in h or "service" in h:
            idx.setdefault("status_raw", i)
    return idx


def _cell_str(vals: tuple[Any, ...], col: int | None) -> str:
    if col is None or col < 1 or col > len(vals):
        return ""
    v = vals[col - 1]
    if v is None:
        return ""
    return str(v).strip()


def _cell_float(vals: tuple[Any, ...], col: int | None) -> float | None:
    s = _cell_str(vals, col)
    if not s:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def _is_battery_row(tech: str, es1: str) -> bool:
    t = (tech or "").upper()
    e = (es1 or "").upper().strip()
    return "BATTERIES" in t or t.endswith("BATTERY") or e in ("MWH", "BAT")


def resolve_latest_860m_generator_xlsx() -> tuple[str, str]:
    """Return (absolute_url, filename) for the newest *_generatorYYYY.xlsx on the index."""
    resp = requests.get(EIA_860M_INDEX, headers={"User-Agent": USER_AGENT}, timeout=120)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")
    best_key: tuple[int, int, str] | None = None
    best_url: str | None = None
    for a in soup.find_all("a", href=True):
        href = a["href"]
        name = href.split("/")[-1].lower()
        m = re.match(r"([a-z]{3})_generator(20\d{2})\.xlsx$", name)
        if not m:
            continue
        mon = _MONTH_ORDER.get(m.group(1), 0)
        year = int(m.group(2))
        if mon == 0:
            continue
        key = (year, mon)
        if best_key is None or key > (best_key[0], best_key[1]):
            best_key = (year, mon, name)
            best_url = href if href.startswith("http") else urljoin(EIA_860M_INDEX, href)
    if not best_url or not best_key:
        raise RuntimeError("could not resolve latest 860M *_generator*.xlsx from index")
    return best_url, best_key[2]


def download_xlsx(url: str) -> bytes:
    r = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=600)
    r.raise_for_status()
    return r.content


def build_battery_rows(wb: openpyxl.workbook.workbook.Workbook) -> list[tuple]:
    handlers: list[tuple[str, Any]] = [
        ("Operating", _operable_status),
        ("Planned", _map_proposed_status),
    ]
    for retired_sheet in ("Retired", "Retired and Canceled", "Retired and Cancelled"):
        if retired_sheet in wb.sheetnames:
            handlers.append((retired_sheet, _retired_status))
            break

    by_id: dict[str, tuple] = {}
    for sheet_name, status_fn in handlers:
        if sheet_name not in wb.sheetnames:
            logger.warning("sheet %s not found — skip", sheet_name)
            continue
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        try:
            header_row = next(it)
        except StopIteration:
            continue
        if not header_row:
            continue
        cmap = _col_map(tuple(header_row))
        need = ("plant_id", "generator_id", "state", "capacity_mw")
        if not all(k in cmap for k in need):
            logger.warning("sheet %s missing required columns %s", sheet_name, need)
            continue
        for raw in it:
            if not raw:
                continue
            row = tuple(raw)
            technology = _cell_str(row, cmap.get("technology"))
            es1 = _cell_str(row, cmap.get("energy_source_1"))
            if not _is_battery_row(technology, es1):
                continue
            plant_id = _cell_str(row, cmap.get("plant_id"))
            gen_id = _cell_str(row, cmap.get("generator_id"))
            if not plant_id or not gen_id:
                continue
            cap_mw = _cell_float(row, cmap.get("capacity_mw"))
            if cap_mw is None or cap_mw <= 0:
                continue
            cap_mwh = _cell_float(row, cmap.get("capacity_mwh"))
            duration = None
            if cap_mwh is not None and cap_mw:
                duration = float(cap_mwh) / float(cap_mw)
            state = _cell_str(row, cmap.get("state")).upper()[:2] or "ZZ"
            lat = _cell_float(row, cmap.get("lat"))
            lon = _cell_float(row, cmap.get("lon"))
            if lat is None or lon is None:
                continue
            plant_name = _cell_str(row, cmap.get("plant_name")) or f"Plant {plant_id}"
            gen_name = _cell_str(row, cmap.get("generator_name"))
            name = gen_name or plant_name
            owner = _cell_str(row, cmap.get("owner")) or None
            st_raw = _cell_str(row, cmap.get("status_raw"))
            status = status_fn(st_raw)
            oid = f"eia-{plant_id}-{gen_id}"
            cod: date | None = None
            if cmap.get("operating_date"):
                cod = _parse_date(row[cmap["operating_date"] - 1])
            ret: date | None = None
            if cmap.get("retirement_date"):
                ret = _parse_date(row[cmap["retirement_date"] - 1])
            iso = iso_for_point(lat, lon)
            try:
                eia_plant_int = int(float(plant_id))
            except ValueError:
                eia_plant_int = None
            by_id[oid] = (
                oid,
                eia_plant_int,
                gen_id or None,
                name,
                owner,
                iso,
                state,
                float(cap_mw),
                float(cap_mwh) if cap_mwh is not None else None,
                float(duration) if duration is not None else None,
                status,
                cod,
                ret,
                lon,
                lat,
            )
    return list(by_id.values())


UPSERT_SQL = """
INSERT INTO battery_assets (
  id, eia_plant_id, eia_generator_id, name, owner, iso, state,
  capacity_mw, capacity_mwh, duration_hours, status, cod_date, retirement_date, geom
) VALUES (
  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
  ST_SetSRID(ST_MakePoint(%s, %s), 4326)
)
ON CONFLICT (id) DO UPDATE SET
  eia_plant_id = EXCLUDED.eia_plant_id,
  eia_generator_id = EXCLUDED.eia_generator_id,
  name = EXCLUDED.name,
  owner = EXCLUDED.owner,
  iso = EXCLUDED.iso,
  state = EXCLUDED.state,
  capacity_mw = EXCLUDED.capacity_mw,
  capacity_mwh = EXCLUDED.capacity_mwh,
  duration_hours = EXCLUDED.duration_hours,
  status = EXCLUDED.status,
  cod_date = EXCLUDED.cod_date,
  retirement_date = EXCLUDED.retirement_date,
  geom = EXCLUDED.geom;
"""


def main() -> int:
    url, fname = resolve_latest_860m_generator_xlsx()
    logger.info("latest EIA 860M generator workbook: %s (%s)", fname, url)
    raw = download_xlsx(url)
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    rows = build_battery_rows(wb)
    wb.close()
    logger.info("parsed %s battery rows (pre-upsert)", len(rows))
    iso_ct = Counter(r[5] for r in rows)
    logger.info("ISO distribution: %s", dict(iso_ct.most_common(12)))

    conn = connect()
    try:
        with conn.cursor() as cur:
            for batch_start in range(0, len(rows), 500):
                batch = rows[batch_start : batch_start + 500]
                cur.executemany(UPSERT_SQL, batch)
                conn.commit()
    finally:
        conn.close()
    logger.info("upserted %s battery_assets rows", len(rows))
    return 0


if __name__ == "__main__":
    sys.exit(main())
